import csv
from openpyxl.worksheet.table import TABLESTYLES
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from string import ascii_uppercase


wb = load_workbook(filename='output.xlsx')
ws = wb.create_sheet("130-001766")
wbr = load_workbook(filename = 'read.xlsx')

sheet_ranges5P = wbr['130-001766-P']
sheet_ranges5S = wbr['130-001766-S']


# initial values
ws['A1'] = "TABLE"
ws['B1'] = "130-001766"
ws['A2'] = "ANGL"
ws['B2'] = 0

# writes the ANGL in correct columns from 0 to 85
for x in range(17):
    row = 402*(x+1) + 2
    deg = 5*(x+1)
    ws['A' + str(row)] = "ANGL"
    ws['B' + str(row)] = 5*(x+1)

# writes WAVE and wavelength on all columns
for x in range(18):
    row2 = 3 + (x)*402
    col = ascii_uppercase[3 + x]
    for y in range(401):
        ws['A' + str(row2 + y)] = 'WAVE'
        ws['B' + str(row2 + y)] = (y+400)/1000

        # iterate through alphabet
        # iterate on row2

        # will write transmittances and reflectances

        # Rs
        ws['C' + str(row2 + y)] = '=1-E' + str(row2 + y)
        # Rp
        ws['D' + str(row2 + y)] = '=1-F' + str(row2 + y)
        # Ts
        ws['E' + str(row2 + y)] = (sheet_ranges5S[str(col) + str(y + 15)].value)/100
        # Tp
        ws['F' + str(row2 + y)] = (sheet_ranges5P[str(col) + str(y + 15)].value)/100



wb.save('output.xlsx')
