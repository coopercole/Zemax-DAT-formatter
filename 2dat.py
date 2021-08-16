from openpyxl.worksheet.table import TABLESTYLES
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from string import ascii_uppercase

# initialize workbook
wb = Workbook()
ws = wb.active
# worksheet title
ws.title = "130-004781"

# title of workbook you want to read
wbr = load_workbook(filename = 'read.xlsx')
# worksheets you want to read from
sheet_ranges1P = wbr['130-004781-P']
sheet_ranges1S = wbr['130-004781-S']

# initial values for dat
ws['A1'] = "TABLE"
# table name
ws['B1'] = "NV3_DUAL_EX_BP_V2_004781"
ws['A2'] = "ANGL"
ws['B2'] = 0

# writes the ANGL in correct columns from 0 to 85 deg
for x in range(17):
    row = 402*(x+1) + 2
    deg = 5*(x+1)
    ws['A' + str(row)] = "ANGL"
    ws['B' + str(row)] = 5*(x+1)

# writes WAVE and wavelength on all columns
for x in range(18):
    row2 = 3 + (x)*402
    col = ascii_uppercase[3 + x]
    for y in range(401):
        ws['A' + str(row2 + y)] = 'WAVE'
        ws['B' + str(row2 + y)] = (y+400)/1000

        # will write transmittances and reflectances from workbook you are reading
        # uses row2, col, and y to read and write correct cells

        # Rs
        ws['C' + str(row2 + y)] = '=1-E' + str(row2 + y)
        ws['C' + str(row2 + y)].number_format = '0.00E+00'
        # Rp
        ws['D' + str(row2 + y)] = '=1-F' + str(row2 + y)
        ws['D' + str(row2 + y)].number_format = '0.00E+00'
        # Ts
        ws['E' + str(row2 + y)] = (sheet_ranges1S[str(col) + str(y + 15)].value)/100
        ws['E' + str(row2 + y)].number_format = '0.00E+00'
        # Tp
        ws['F' + str(row2 + y)] = (sheet_ranges1P[str(col) + str(y + 15)].value)/100
        ws['F' + str(row2 + y)].number_format = '0.00E+00'

# name of workbook to save to
wb.save('output.xlsx')
